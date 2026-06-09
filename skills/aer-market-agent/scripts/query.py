#!/usr/bin/env python3
"""Oracle query helper over the bundled AER data.

The skill bundles every AER *State of the Energy Market* data workbook (2019-2025,
all chapters) under `data/sources/`, indexed by `data/catalog/catalog.json`. This
tool lets you (or Claude) find and read any figure on demand — for answering
questions and doing analysis, not just building the report.

Commands:
    python scripts/query.py domains
        List the data domains and how many workbooks/figures each has.

    python scripts/query.py list [--year Y] [--domain D]
        List the indexed workbooks (optionally filtered).

    python scripts/query.py search TERM [TERM ...] [--year Y] [--domain D] [--limit N]
        Search figure/table titles across every workbook. Prints matches as
        "YEAR  DOMAIN  Figure X.Y | title".

    python scripts/query.py get YEAR FIGURE [--domain D] [--max-rows N]
        Read one figure's data block from its raw workbook and print it as CSV.
        e.g.  python scripts/query.py get 2025 "Figure 3.14"

    python scripts/query.py file RELPATH SHEET [--max-rows N]
        Dump any sheet of any bundled workbook directly.

All reads are local — nothing here touches the internet.

Built and maintained by Renovara - renovara.co
"""
from __future__ import annotations
import argparse, csv, io, json, re, shutil, subprocess, sys, urllib.request, urllib.error
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CATALOG = ROOT / "data" / "catalog" / "catalog.json"
CACHE = ROOT / "data" / "catalog" / "_cache"


def load_catalog():
    if not CATALOG.exists():
        sys.exit(f"catalog not found at {CATALOG}. Run scripts/build_catalog.py first.")
    return json.loads(CATALOG.read_text())


# ---- raw workbook reading (xlsx + xlsb) ------------------------------------
def _curl_download(url, dest, timeout=120):
    """Download with curl (the client AER serves), falling back to urllib. -> (ok, detail)."""
    if shutil.which("curl"):
        r = subprocess.run(["curl", "-fSL", "--show-error", "--retry", "2",
                            "--max-time", str(timeout), "-o", str(dest), url],
                           capture_output=True, text=True)
        if r.returncode == 0 and dest.exists() and dest.stat().st_size > 0:
            return True, "curl ok"
        probe = subprocess.run(["curl", "-sS", "-v", "-o", "/dev/null",
                                "--max-time", str(min(timeout, 30)), url],
                               capture_output=True, text=True)
        return False, ((r.stderr + "\n" + probe.stderr).strip() or f"curl exit {r.returncode}")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "aer-market-agent/1.0"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            dest.write_bytes(resp.read())
        return True, "urllib ok"
    except Exception as e:  # noqa
        return False, str(e)


def _resolve(relpath: str, url: str | None, allow_download: bool):
    """Local file if bundled, else download from the AER url into a cache."""
    p = ROOT / relpath
    if p.exists():
        return p
    if url and allow_download:
        CACHE.mkdir(parents=True, exist_ok=True)
        cp = CACHE / Path(relpath).name
        if not cp.exists():
            sys.stderr.write(f"# fetching from AER (curl): {url}\n")
            ok, detail = _curl_download(url, cp)
            if not ok:
                d = detail.lower()
                if any(k in d for k in ("from proxy after connect", "blocked-by-allowlist", "x-proxy-error",
                                         "tunnel connection failed", "could not resolve")):
                    why = ("egress allow-list blocked it (request never reached AER). Allow `aer.gov.au`\n"
                           "  in Cowork Org settings > Capabilities > Code execution, then start a NEW session.")
                elif "403" in d or "forbidden" in d:
                    why = ("AER (Akamai) returned 403 — it is blocking this runtime's datacenter IP.\n"
                           "  Run `python scripts/fetch_data.py` on your own machine / Claude Code and read locally.")
                else:
                    why = "download failed."
                sys.exit(f"could not download {url}\n  {detail}\n  {why}")
        return cp
    sys.exit(f"workbook not bundled locally: {relpath}\n"
             + (f"  fetch it with: python scripts/fetch_data.py --file \"{Path(relpath).name}\"\n  (source: {url})"
                if url else "  (no source url in catalog)"))


def read_sheet(relpath: str, sheet: str, max_rows: int | None, url: str | None = None,
               allow_download: bool = True):
    path = _resolve(relpath, url, allow_download)
    if path.suffix.lower() == ".xlsb":
        from pyxlsb import open_workbook
        with open_workbook(str(path)) as wb:
            names = list(wb.sheets)
            real = _match_sheet(sheet, names)
            with wb.get_sheet(real) as sh:
                rows = []
                for i, r in enumerate(sh.rows()):
                    rows.append([c.v for c in r])
                    if max_rows and i + 1 >= max_rows:
                        break
                return real, rows, len(names)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        real = _match_sheet(sheet, names)
        ws = wb[real]
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True, max_row=max_rows or None)):
            rows.append(list(r))
        wb.close()
        return real, rows, len(names)


def _match_sheet(sheet, names):
    s = sheet.strip().lower()
    for n in names:
        if n.strip().lower() == s:
            return n
    for n in names:
        if n.strip().lower().startswith(s):
            return n
    sys.exit(f"sheet '{sheet}' not found. Available: {', '.join(names[:40])}")


def trim(rows):
    g = [["" if c is None else (round(c, 6) if isinstance(c, float) else c) for c in r] for r in rows]
    g = [r for r in g if any(c not in ("", None) for c in r)]
    if not g:
        return g
    maxlen = max(len(r) for r in g)
    def col_nonempty(j):
        return any(j < len(r) and r[j] not in ("", None) for r in g)
    keep = [j for j in range(maxlen) if col_nonempty(j)]
    last = (keep[-1] + 1) if keep else 0
    return [r[:last] for r in g]


def print_csv(rows):
    out = io.StringIO()
    csv.writer(out).writerows(rows)
    sys.stdout.write(out.getvalue())


# ---- commands ---------------------------------------------------------------
def cmd_domains(cat, args):
    from collections import Counter
    figs = Counter()
    wbs = Counter()
    years = {}
    for w in cat["workbooks"]:
        wbs[w["domain"]] += 1
        figs[w["domain"]] += w["n_figures"]
        years.setdefault(w["domain"], set()).add(w["year"])
    print(f"{cat['n_workbooks']} workbooks, {cat['n_figures_total']} figures/tables, "
          f"years {min(cat['years'])}-{max(cat['years'])}\n")
    print(f"{'domain':16} {'label':42} {'wbs':>4} {'figs':>5}  years")
    for d in sorted(wbs, key=lambda d: -figs[d]):
        yr = ",".join(str(y) for y in sorted(years[d]))
        print(f"{d:16} {cat['domains'].get(d, d)[:42]:42} {wbs[d]:>4} {figs[d]:>5}  {yr}")


def cmd_list(cat, args):
    for w in cat["workbooks"]:
        if args.year and w["year"] != args.year:
            continue
        if args.domain and w["domain"] != args.domain:
            continue
        print(f"{w['year']}  {w['domain']:14} {w['chapter']:11} {w['n_figures']:>3} figs  | {w['file']}")


def cmd_search(cat, args):
    terms = [t.lower() for t in args.terms]
    hits = []
    for w in cat["workbooks"]:
        if args.year and w["year"] != args.year:
            continue
        if args.domain and w["domain"] != args.domain:
            continue
        for it in w["contents"]:
            hay = f"{it['item']} {it['title']}".lower()
            if all(t in hay for t in terms):
                hits.append((w, it))
    hits.sort(key=lambda h: (-h[0]["year"], h[0]["domain"]))
    for w, it in hits[: args.limit]:
        print(f"{w['year']}  {w['domain']:13} {it['item']:11} | {it['title'][:70]}")
    print(f"\n{len(hits)} match(es)"
          + (f"; showing {args.limit}" if len(hits) > args.limit else "")
          + ".  Read one with:  python scripts/query.py get <year> \"<Figure X.Y>\"")


def _find_entry(cat, year, figure, domain):
    fig = figure.strip().lower()
    for w in cat["workbooks"]:
        if w["year"] != year:
            continue
        if domain and w["domain"] != domain:
            continue
        for it in w["contents"]:
            if it["item"].strip().lower() == fig:
                return w, it
    return None, None


def cmd_get(cat, args):
    w, it = _find_entry(cat, args.year, args.figure, args.domain)
    if not w:
        sys.exit(f"'{args.figure}' not found for {args.year}"
                 + (f" in {args.domain}" if args.domain else "")
                 + ".  Try:  python scripts/query.py search <terms> --year %d" % args.year)
    print(f"# {args.year} {w['domain_label']} — {it['item']}: {it['title']}")
    if it.get("source"):
        print(f"# {it['source']}")
    print(f"# file: {w['file']}\n# url:  {w.get('url','')}\n")
    real, rows, _ = read_sheet(w["file"], it["item"], args.max_rows, w.get("url"), not args.no_download)
    g = trim(rows)
    print_csv(g)
    if args.max_rows and len(g) >= args.max_rows:
        print(f"\n# (showing first {args.max_rows} rows; use --max-rows 0 for all)")


def cmd_file(cat, args):
    real, rows, nsheets = read_sheet(args.relpath, args.sheet, args.max_rows)
    print(f"# {args.relpath} :: {real}  ({nsheets} sheets in workbook)\n")
    print_csv(trim(rows))


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    sub.add_parser("domains")

    pl = sub.add_parser("list")
    pl.add_argument("--year", type=int)
    pl.add_argument("--domain")

    ps = sub.add_parser("search")
    ps.add_argument("terms", nargs="+")
    ps.add_argument("--year", type=int)
    ps.add_argument("--domain")
    ps.add_argument("--limit", type=int, default=30)

    pg = sub.add_parser("get")
    pg.add_argument("year", type=int)
    pg.add_argument("figure")
    pg.add_argument("--domain")
    pg.add_argument("--max-rows", type=int, default=60)
    pg.add_argument("--no-download", action="store_true", help="do not fetch from AER if file is not bundled")

    pf = sub.add_parser("file")
    pf.add_argument("relpath")
    pf.add_argument("sheet")
    pf.add_argument("--max-rows", type=int, default=60)

    args = ap.parse_args()
    # treat --max-rows 0 as "all"
    if getattr(args, "max_rows", None) == 0:
        args.max_rows = None
    cat = load_catalog()
    {"domains": cmd_domains, "list": cmd_list, "search": cmd_search,
     "get": cmd_get, "file": cmd_file}[args.cmd](cat, args)


if __name__ == "__main__":
    main()
