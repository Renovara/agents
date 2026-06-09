#!/usr/bin/env python3
"""Build data/catalog/catalog.json — the index of AER State of the Energy Market
workbooks (2019-2025, all chapters).

Driven by data/catalog/source_urls.json (filename -> AER download URL). For each
workbook it reads the sheet list and each figure/table's title (+ source line) from
the local copy under data/sources/<year>/ if present, otherwise it downloads the
file from its AER URL into data/catalog/_cache/ and reads that. So the full catalog
regenerates whether or not the raw files are bundled.

    python scripts/build_catalog.py            # local file if present, else download
    python scripts/build_catalog.py --local    # only index locally-present files

Built and maintained by Renovara - renovara.co
"""
from __future__ import annotations
import argparse, json, re, sys, urllib.request
from pathlib import Path
from collections import Counter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "sources"
OUT = ROOT / "data" / "catalog"
CACHE = OUT / "_cache"

DOMAIN_RULES = [
    ("electricity networks", "networks"),
    ("national electricity market", "wholesale-nem"),
    ("national electricity market", "wholesale-nem"),
    ("electricity market in transition", "transition"),
    ("energy system in transition", "transition"),
    ("energy market in transition", "transition"),
    ("retail energy markets", "retail"),
    ("gas markets", "gas-markets"),
    ("regulated gas networks", "gas-pipelines"),
    ("regulated gas pipelines", "gas-pipelines"),
    ("gas pipelines", "gas-pipelines"),
    ("market overview", "overview"),
]
DOMAIN_LABEL = {
    "networks": "Electricity networks",
    "wholesale-nem": "National Electricity Market (wholesale)",
    "transition": "Energy market in transition",
    "retail": "Retail energy markets",
    "gas-markets": "Gas markets (eastern Australia)",
    "gas-pipelines": "Gas pipelines / regulated gas networks",
    "overview": "Market overview",
    "other": "Other",
}
DATA_SHEET = re.compile(r"^(Figure|Table|Chart|Map)\s", re.I)


def parse_filename(name: str):
    yr = re.search(r"\b(20\d{2})\b", name)
    year = int(yr.group(1)) if yr else 0
    rest = re.sub(r"^Data\s*-\s*State of the [Ee]nergy [Mm]arket\s*20\d{2}( update)?\s*-\s*", "", name)
    rest = re.sub(r"\.(xlsx|xlsb)$", "", rest, flags=re.I)
    rest = re.sub(r"\s*-\s*amended.*$", "", rest, flags=re.I)
    rest = re.sub(r"_\d+$", "", rest)
    m = re.match(r"Chapter\s*(\d+)\s*-?\s*(.*)$", rest)
    if m:
        chap_label, chap_title = f"Chapter {m.group(1)}", m.group(2).strip(" -")
    else:
        chap_label = chap_title = rest.strip()
    domain = next((d for kw, d in DOMAIN_RULES if kw in chap_title.lower()), "other")
    return year, chap_label, chap_title, domain


def _title_source(rows):
    title = source = ""
    for row in rows:
        for c in row:
            if isinstance(c, str) and c.strip():
                s = c.strip()
                if not title and DATA_SHEET.match(s):
                    title = re.sub(r"^(Figure|Table|Chart|Map)\s*[\d.]+\s*[–—:-]?\s*", "", s).strip() or s
                elif not title:
                    title = s
                if not source and s.lower().startswith("source:"):
                    source = s
    return title, source


def index_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    items = []
    for n in names:
        if DATA_SHEET.match(n.strip()):
            rows = [r for r in wb[n].iter_rows(values_only=True, max_row=5)]
            t, s = _title_source(rows)
            items.append({"item": n.strip(), "title": t[:200], "source": s[:240]})
    wb.close()
    return names, items


def index_xlsb(path):
    from pyxlsb import open_workbook
    items = []
    with open_workbook(str(path)) as wb:
        names = list(wb.sheets)
        for n in names:
            if DATA_SHEET.match(n.strip()):
                with wb.get_sheet(n) as sh:
                    rows = []
                    for i, r in enumerate(sh.rows()):
                        rows.append([c.v for c in r])
                        if i >= 4:
                            break
                t, s = _title_source(rows)
                items.append({"item": n.strip(), "title": t[:200], "source": s[:240]})
    return names, items


def resolve(basename, year, url, local_only):
    """Return a local path to the workbook, downloading from url if needed."""
    p = SRC / str(year) / basename
    if p.exists():
        return p
    if local_only or not url:
        return None
    CACHE.mkdir(parents=True, exist_ok=True)
    cp = CACHE / basename
    if not cp.exists():
        print(f"    downloading {basename[:50]} ...")
        urllib.request.urlretrieve(url, cp)
    return cp


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--local", action="store_true", help="only index files present under data/sources/")
    args = ap.parse_args()

    urls = {}
    uf = OUT / "source_urls.json"
    if uf.exists():
        urls = json.loads(uf.read_text())
    # the set to index: every filename in the url map, plus any extra local files
    names = set(urls)
    if SRC.exists():
        for p in SRC.glob("[0-9][0-9][0-9][0-9]/*.xls*"):
            if not p.name.startswith("~$"):
                names.add(p.name)

    entries = []
    for basename in sorted(names):
        year, chap_label, chap_title, domain = parse_filename(basename)
        url = urls.get(basename, "")
        path = resolve(basename, year, url, args.local)
        if path is None:
            print(f"  SKIP (no local file, no url) {basename[:50]}")
            continue
        try:
            sheets, items = (index_xlsb if path.suffix.lower() == ".xlsb" else index_xlsx)(path)
        except Exception as e:  # noqa
            print(f"  ERROR {basename}: {e}")
            continue
        entries.append({
            "year": year, "domain": domain, "domain_label": DOMAIN_LABEL.get(domain, domain),
            "chapter": chap_label, "chapter_title": chap_title,
            "file": f"data/sources/{year}/{basename}", "url": url,
            "n_sheets": len(sheets), "n_figures": len(items),
            "sheets": sheets, "contents": items,
        })
        print(f"  {year} {domain:14} {chap_label:11} {len(items):3} figs  <- {basename[:44]}")

    entries.sort(key=lambda e: (e["year"], e["domain"], e["chapter"]))
    catalog = {
        "description": "Index of AER State of the Energy Market data workbooks (2019-2025, all chapters). "
                       "Each entry has 'url' (direct AER download) and 'file' (local path if bundled). "
                       "scripts/query.py reads a figure on demand, using the local file or downloading from 'url'.",
        "domains": DOMAIN_LABEL,
        "years": sorted({e["year"] for e in entries}),
        "n_workbooks": len(entries),
        "n_figures_total": sum(e["n_figures"] for e in entries),
        "workbooks": entries,
    }
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "catalog.json").write_text(json.dumps(catalog, indent=2, default=str))
    print(f"\nIndexed {len(entries)} workbooks, {catalog['n_figures_total']} figures -> {OUT/'catalog.json'}")
    print("by domain:", dict(Counter(e["domain"] for e in entries)))


if __name__ == "__main__":
    main()
