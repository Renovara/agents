#!/usr/bin/env python3
"""Rebuild the data-derived fields of dmo_dataset.json from local source files.

The small source files (demand / reference-price CSVs) ship bundled under
`data/sources/`. The large Chapter 3 networks workbook is NOT bundled in the
published skill - if it is missing it is fetched from the AER on first use via
the catalog URL (see DATA_SOURCES.md). The pipeline is:

    data/sources/*  ->  parse_sources.py  ->  data/dmo_dataset.json  ->  build_report.py

Run it after dropping a newer AER workbook / CSV into data/sources/ (see
DATA_SOURCES.md for the exact files and where they come from):

    python scripts/parse_sources.py            # rebuild + write dataset
    python scripts/parse_sources.py --check     # rebuild + compare, don't write

The series pulled from local files:
  - longview.revenue          <- SotEM Ch3 workbook, Figure 3.11 ("Actual revenue")
  - usage_per_customer.values <- SotEM Ch3 workbook, Figure 3.31 (all metered customers)
  - tariff_shift (9% -> 37%)  <- SotEM Ch3 workbook, Figure 3.9 ("Average" row)
  - demand_avg / demand_peak  <- AER demand & energy CSVs (NEM consumption + seasonal peak)
  - determinations / residential reference prices <- DMO reference-price CSV

Everything else in the dataset (cost-stack proportions, worked example, building
blocks, references, narrative copy) is curated and verified by hand - it is NOT
overwritten here. Always sanity-check the printed summary before publishing.

Built and maintained by Renovara - renovara.co
"""
from __future__ import annotations
import argparse, csv, json, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "sources"
DATASET = ROOT / "data" / "dmo_dataset.json"

WORKBOOK = SRC / "2025" / "Data - State of the energy market 2025 - Chapter 3 - Electricity networks.xlsx"
CONSUMPTION_CSV = SRC / "demand" / "AER-NEM-annual-consumption.csv"
PEAK_CSV = SRC / "demand" / "AER-NEM-seasonal-peak.csv"
PRICES_CSV = SRC / "dmo-reference-prices-2019-2027.csv"

# Financial years shown on the demand slide (DMO 1 -> DMO 6 windows).
DEMAND_FYS = ["2019-20", "2020-21", "2021-22", "2022-23", "2023-24", "2024-25"]
HOURS_PER_YEAR = 8.76  # 8760 h / 1000 -> average GW from annual TWh


def _require(path: Path) -> None:
    if not path.exists():
        sys.exit(f"ERROR: missing local source file: {path}\n"
                 f"See DATA_SOURCES.md for what belongs in {SRC}/.")


def _catalog_url_for(rel_path: str) -> str | None:
    """Look up a workbook's AER download URL by its `file` path in the catalog."""
    cat_file = ROOT / "data" / "catalog" / "catalog.json"
    if not cat_file.exists():
        return None
    for w in json.loads(cat_file.read_text()).get("workbooks", []):
        if w.get("file") == rel_path:
            return w.get("url")
    return None


def _ensure_workbook() -> None:
    """Ensure the Ch3 networks workbook is present, fetching it on demand.

    The workbook is large (~4.7 MB) and is deliberately NOT shipped in the
    published skill bundle; it is pulled from the AER on first use. Reuses
    scripts/fetch_data.py's downloader so the curl/urllib + WAF/egress handling
    stays in one place.
    """
    if WORKBOOK.exists():
        return
    rel = WORKBOOK.relative_to(ROOT).as_posix()
    url = _catalog_url_for(rel)
    if not url:
        sys.exit(f"ERROR: missing {WORKBOOK}\n"
                 f"and no catalog URL to fetch it from. See DATA_SOURCES.md.")
    print(f"Chapter 3 workbook not bundled; fetching from AER:\n  {url}")
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import fetch_data
    ok, how, detail = fetch_data.download(url, WORKBOOK, timeout=120)
    if not ok:
        cause = fetch_data.classify(detail)
        hint = ""
        if cause == "egress":
            hint = ("\nEgress allow-list blocked it (never reached AER). On Cowork, add "
                    "`aer.gov.au` and `www.aer.gov.au` under Code execution, then a NEW session.")
        elif cause == "akamai":
            hint = ("\nReached AER but got HTTP 403 from its bot/WAF. Run on your own machine / "
                    "Claude Code where the URL serves fine, then re-run parse_sources.")
        sys.exit(f"ERROR: could not fetch workbook ({how}): {detail}{hint}")
    print(f"  + {detail/1e6:.2f} MB ({how}) -> {WORKBOOK.relative_to(ROOT)}")


def load_workbook():
    try:
        import openpyxl  # noqa
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")
    import openpyxl
    return openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)


def _row_starting_with(ws, label):
    """Return the list of cell values for the first row whose 2nd-ish cell == label."""
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if isinstance(c, str) and c.strip() == label:
                return list(row)
    return None


def parse_longview(wb):
    """Figure 3.11 'Actual revenue' -> real $bn series (2006 -> latest actual)."""
    ws = wb["Figure 3.11"]
    row = _row_starting_with(ws, "Actual revenue")
    vals = [round(v / 1e9, 1) for v in row if isinstance(v, (int, float))]
    return vals


def parse_usage(wb):
    """Figure 3.31 'All metered distribution customers' -> MWh/customer series."""
    ws = wb["Figure 3.31"]
    row = _row_starting_with(ws, "All metered distribution customers (LHS)")
    vals = [round(v, 3) for v in row if isinstance(v, (int, float))]
    return vals


def parse_tariff_avg(wb):
    """Figure 3.9 'Average' row -> first & last share of customers on cost-reflective tariffs."""
    ws = wb["Figure 3.9"]
    # header row holds the years; the 'Average' row holds the shares
    years, avg = None, None
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if any(isinstance(c, (int, float)) and 2015 <= c <= 2035 for c in cells) and years is None:
            years = [int(c) for c in cells if isinstance(c, (int, float))]
        if any(isinstance(c, str) and c.strip() == "Average" for c in cells):
            avg = [c for c in cells if isinstance(c, (int, float))]
    first, last = avg[0], avg[-1]
    y0, y1 = years[0], years[-1]
    return {
        "first_pct": round(first * 100),
        "last_pct": round(last * 100),
        "first_year": y0,
        "last_year": y1,
    }


def _read_csv(path):
    with open(path, newline="") as f:
        return list(csv.reader(f))


def parse_demand_avg():
    rows = _read_csv(CONSUMPTION_CSV)
    header = rows[0]
    ni = header.index("NEM (Terawatt hours)")
    by_fy = {r[0]: r[ni] for r in rows[1:] if r and r[0]}
    return [round(float(by_fy[fy]) / HOURS_PER_YEAR, 1) for fy in DEMAND_FYS]


def parse_demand_peak():
    """Max of summer/winter NEM peak (MW) within each financial year -> GW."""
    rows = _read_csv(PEAK_CSV)
    summer, winter = {}, {}
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        period = r[0].strip()
        if period.startswith("Summer"):  # e.g. 'Summer 2019/20'
            yr = period.split()[1].split("/")[0]
            if r[1]:
                summer[yr] = float(r[1])
        elif period.startswith("Winter"):  # e.g. 'Winter 2020'
            yr = period.split()[1]
            if len(r) > 2 and r[2]:
                winter[yr] = float(r[2])
    out = []
    for fy in DEMAND_FYS:
        start = fy.split("-")[0]               # '2019'
        end = str(int(start) + 1)              # '2020'  (winter falls in calendar end year)
        cand = [v for v in (summer.get(start), winter.get(end)) if v]
        out.append(round(max(cand) / 1000, 1))
    return out


def parse_prices():
    """DMO reference-price CSV -> determinations list keyed for the dataset."""
    rows = _read_csv(PRICES_CSV)
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}
    recs = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        recs.append({
            "determination": r[idx["determination"]],
            "financial_year": r[idx["financial_year"]],
            "region": r[idx["region"]],
            "zone": r[idx["zone"]],
            "residential_flat_aud": int(r[idx["residential_flat_no_cl_aud"]]) if r[idx["residential_flat_no_cl_aud"]] else None,
        })
    return recs


def main():
    ap = argparse.ArgumentParser(description="Rebuild dmo_dataset.json from local source files.")
    ap.add_argument("--check", action="store_true", help="compare against current dataset, do not write")
    args = ap.parse_args()

    _ensure_workbook()
    for p in (CONSUMPTION_CSV, PEAK_CSV, PRICES_CSV):
        _require(p)

    wb = load_workbook()
    derived = {
        "longview.revenue": parse_longview(wb),
        "usage_per_customer.values": parse_usage(wb),
        "tariff_avg": parse_tariff_avg(wb),
        "demand_avg": parse_demand_avg(),
        "demand_peak": parse_demand_peak(),
        "prices": parse_prices(),
    }

    print("Parsed from local files in data/sources/:")
    print(f"  longview.revenue       {len(derived['longview.revenue'])} pts, "
          f"{derived['longview.revenue'][0]} -> {derived['longview.revenue'][-1]} ($bn, real)")
    print(f"  usage_per_customer     {len(derived['usage_per_customer.values'])} pts, "
          f"{derived['usage_per_customer.values'][0]} -> {derived['usage_per_customer.values'][-1]} (MWh)")
    t = derived["tariff_avg"]
    print(f"  cost-reflective tariffs {t['first_pct']}% ({t['first_year']}) -> {t['last_pct']}% ({t['last_year']})")
    print(f"  demand_avg (GW)        {derived['demand_avg']}")
    print(f"  demand_peak (GW)       {derived['demand_peak']}")
    print(f"  reference prices        {len(derived['prices'])} zone-year records")

    ds = json.loads(DATASET.read_text())

    # --- apply to dataset (in-memory) ---
    new = json.loads(json.dumps(ds))  # deep copy
    new["longview"]["revenue"] = derived["longview.revenue"]
    new["usage_per_customer"]["values"] = derived["usage_per_customer.values"]
    new["tariff_shift"]["tariff_value"] = f"{t['first_pct']}% -> {t['last_pct']}%"
    new["tariff_shift"]["tariff_detail"] = (
        f"residential customers on cost-reflective tariffs, {t['first_year']}->{t['last_year']}")
    new["demand_avg"] = derived["demand_avg"]
    new["demand_peak"] = derived["demand_peak"]

    if args.check:
        diffs = []
        for path, val in [
            ("longview.revenue", new["longview"]["revenue"]),
            ("usage_per_customer.values", new["usage_per_customer"]["values"]),
            ("demand_avg", new["demand_avg"]),
            ("demand_peak", new["demand_peak"]),
            ("tariff_shift.tariff_value", new["tariff_shift"]["tariff_value"]),
        ]:
            cur = ds
            for k in path.split("."):
                cur = cur[k]
            if cur != val:
                diffs.append((path, cur, val))
        if diffs:
            print("\nDIFFERENCES vs current dataset:")
            for p, a, b in diffs:
                print(f"  {p}\n    current: {a}\n    rebuilt: {b}")
            sys.exit(1)
        print("\nCheck passed: rebuilt series match the current dataset exactly.")
        return

    DATASET.write_text(json.dumps(new, indent=2))
    print(f"\nWrote {DATASET.relative_to(ROOT)} from local sources.")


if __name__ == "__main__":
    main()
