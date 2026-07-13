#!/usr/bin/env python3
"""Query the 2026 ISP document catalog and read data from workbooks/zips.

The oracle for the isp-2026-agent skill. Everything reads the authoritative
AEMO file — nothing is pre-extracted. If a file isn't local yet, `sheets`,
`get` and `zip` auto-download it from its catalog url (same network rules as
fetch_data.py: may need to run outside a sandbox).

    python scripts/query.py categories                 # what's covered
    python scripts/query.py search chart data          # find documents by keyword
    python scripts/query.py search demand traces NSW --category demand-traces-regional
    python scripts/query.py info 2026-isp-chart-data   # metadata + url + local path
    python scripts/query.py sheets 2026-isp-chart-data # list workbook sheets
    python scripts/query.py get 2026-isp-chart-data "Figure 1" --max-rows 40
    python scripts/query.py zip 2026-isp-generation-and-storage-outlook   # list zip contents
    python scripts/query.py zip <id> --extract "*.csv"                    # extract members
    python scripts/query.py pages                      # AEMO pages (methodology, consultations)

`get` prints the sheet's used range as CSV. PDFs aren't parsed here — fetch
them and read directly (they're normal PDFs).

Requires: openpyxl (pip install openpyxl) for sheets/get.

Built and maintained by Renovara - renovara.co
"""
from __future__ import annotations
import argparse, csv, fnmatch, io, json, sys, zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CATALOG = ROOT / "data" / "catalog" / "catalog.json"


def load():
    if not CATALOG.exists():
        sys.exit(f"catalog not found at {CATALOG}")
    return json.loads(CATALOG.read_text())


def find(cat, doc_id):
    for d in cat["documents"]:
        if d["id"] == doc_id:
            return d
    sys.exit(f"no document with id '{doc_id}' — try: python scripts/query.py search <terms>")


def ensure_local(d):
    dest = ROOT / d["file"]
    if dest.exists():
        return dest
    print(f"[fetching {d['id']} from AEMO ...]", file=sys.stderr)
    sys.path.insert(0, str(Path(__file__).parent))
    from fetch_data import download, classify
    ok, how, detail = download(d["url"], dest, 600)
    if not ok:
        cause = classify(str(detail))
        msg = f"download failed ({how}): {detail}"
        if cause == "egress":
            msg += ("\n\nThe sandbox egress allow-list blocked this. Run the fetch on your own "
                    "machine / Claude Code:\n  python scripts/fetch_data.py --id " + d["id"] +
                    "\nor allow `aemo.com.au` in Cowork org settings (see fetch_data.py --help).")
        if cause == "waf":
            msg += ("\n\nAEMO's WAF blocked this runtime's IP. Run the same fetch on your own "
                    "machine and place the file at " + d["file"])
        sys.exit(msg)
    return dest


def cmd_categories(cat, _):
    counts, sizes = {}, {}
    for d in cat["documents"]:
        counts[d["category"]] = counts.get(d["category"], 0) + 1
    print(f"{len(cat['documents'])} documents. Scenarios: {', '.join(cat['scenarios'])}")
    print(f"{'category':28} {'#':>4}  local")
    for c in sorted(counts):
        local = sum(1 for d in cat["documents"]
                    if d["category"] == c and (ROOT / d["file"]).exists())
        print(f"{c:28} {counts[c]:>4}  {local}")


def cmd_search(cat, args):
    terms = [t.lower() for t in args.terms]
    hits = []
    for d in cat["documents"]:
        if args.category and d["category"] != args.category:
            continue
        hay = " ".join([d["title"], d["id"], d["category"], d["type"], d.get("notes", "")]).lower()
        if all(t in hay for t in terms):
            hits.append(d)
    if not hits:
        sys.exit("no matches.")
    for d in hits[:args.max_hits]:
        loc = "local" if (ROOT / d["file"]).exists() else "remote"
        print(f"{d['id']:60} [{d['category']}] {d['type']} {d.get('size','')} ({loc})")
        if d.get("notes"):
            print(f"    {d['notes']}")
    if len(hits) > args.max_hits:
        print(f"... and {len(hits) - args.max_hits} more (raise --max-hits)")


def cmd_info(cat, args):
    d = find(cat, args.id)
    for k in ("id", "title", "category", "type", "date", "size", "notes", "url", "file"):
        if k in d:
            print(f"{k:9} {d[k]}")
    print(f"{'local':9} {(ROOT / d['file']).exists()}")


def open_wb(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl required: pip install openpyxl")
    return openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)


def cmd_sheets(cat, args):
    d = find(cat, args.id)
    if d["type"] not in ("xlsx", "xlsm"):
        sys.exit(f"'{d['id']}' is {d['type']}, not a workbook.")
    wb = open_wb(ensure_local(d))
    for ws in wb.worksheets:
        print(f"{ws.title:40} {ws.max_row:>7} rows x {ws.max_column:>3} cols")


def cmd_get(cat, args):
    d = find(cat, args.id)
    if d["type"] not in ("xlsx", "xlsm"):
        sys.exit(f"'{d['id']}' is {d['type']}, not a workbook. For zips use `zip`, PDFs read directly.")
    wb = open_wb(ensure_local(d))
    names = [ws.title for ws in wb.worksheets]
    match = [n for n in names if n.lower() == args.sheet.lower()] or \
            [n for n in names if args.sheet.lower() in n.lower()]
    if not match:
        sys.exit(f"no sheet matching '{args.sheet}'. Sheets: {', '.join(names)}")
    ws = wb[match[0]]
    print(f"# {d['id']} :: {match[0]}", file=sys.stderr)
    w = csv.writer(sys.stdout)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if args.max_rows and i >= args.max_rows:
            print(f"# ... truncated at {args.max_rows} rows (use --max-rows 0 for all)",
                  file=sys.stderr)
            break
        w.writerow(["" if v is None else v for v in row])


def cmd_zip(cat, args):
    d = find(cat, args.id)
    if d["type"] != "zip":
        sys.exit(f"'{d['id']}' is {d['type']}, not a zip.")
    path = ensure_local(d)
    with zipfile.ZipFile(path) as z:
        if args.extract:
            outdir = path.parent / (path.stem + "-extracted")
            n = 0
            for m in z.namelist():
                if fnmatch.fnmatch(m, args.extract):
                    z.extract(m, outdir)
                    n += 1
            print(f"extracted {n} member(s) to {outdir}")
        else:
            for m in z.infolist():
                print(f"{m.file_size/1e6:9.2f} MB  {m.filename}")


def cmd_pages(cat, _):
    for k, v in cat["pages"].items():
        print(f"{k:40} {v}")


def main():
    ap = argparse.ArgumentParser(description="Query 2026 ISP documents.",
                                 formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("categories")
    sp = sub.add_parser("search")
    sp.add_argument("terms", nargs="+")
    sp.add_argument("--category")
    sp.add_argument("--max-hits", type=int, default=25)
    sp = sub.add_parser("info"); sp.add_argument("id")
    sp = sub.add_parser("sheets"); sp.add_argument("id")
    sp = sub.add_parser("get")
    sp.add_argument("id"); sp.add_argument("sheet")
    sp.add_argument("--max-rows", type=int, default=60,
                    help="0 = all rows (default 60)")
    sp = sub.add_parser("zip")
    sp.add_argument("id"); sp.add_argument("--extract", metavar="GLOB")
    sub.add_parser("pages")
    args = ap.parse_args()
    cat = load()
    {"categories": cmd_categories, "search": cmd_search, "info": cmd_info,
     "sheets": cmd_sheets, "get": cmd_get, "zip": cmd_zip, "pages": cmd_pages}[args.cmd](cat, args)


if __name__ == "__main__":
    main()
